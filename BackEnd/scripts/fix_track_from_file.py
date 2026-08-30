#!/usr/bin/env python3
"""Fix trackpoints (velocity/width/depth) by re-parsing an uploaded Excel/CSV file and updating the sqlite DB.

Usage:
  python fix_track_from_file.py <file_path> <trackid>

This script matches points by gpstime+lon+lat and updates numeric fields when missing.
"""
import sys
import sqlite3
import math
from pathlib import Path
import csv
from openpyxl import load_workbook


def normalize_headers(headers):
    import re
    def norm(v):
        return re.sub(r"[\s\-_/\\.:：，,。()（）\[\]{}]+", "", str(v).strip().lower())
    return [norm(h) for h in headers]


def parse_xlsx(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    nheaders = normalize_headers(headers)
    # find indexes
    idx = {}
    for i, h in enumerate(nheaders):
        if h in ('gpstime','time','gpstime','gps时间'.lower()) or 'time' in h:
            idx.setdefault('gpstime', i)
        if 'lon' in h or '经度' in h:
            idx.setdefault('lon', i)
        if 'lat' in h or '纬度' in h:
            idx.setdefault('lat', i)
        if 'speed' in h or 'velocity' in h or '速度' in h:
            idx.setdefault('velocity', i)
        if 'width' in h or '幅宽' in h:
            idx.setdefault('width', i)
        if 'depthstandard' in h or '标准值' in h or '深度标准' in h or '耕深标准' in h:
            idx.setdefault('depthstandard', i)
        elif 'depth' in h or '耕深' in h or ('深度' in h and '标准' not in h):
            idx.setdefault('depth', i)

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        def v(k):
            i = idx.get(k)
            return row[i] if i is not None and i < len(row) else None
        rows.append({
            'gpstime': v('gpstime'),
            'lon': v('lon'),
            'lat': v('lat'),
            'velocity': v('velocity'),
            'width': v('width'),
            'depth': v('depth'),
            'depthstandard': v('depthstandard'),
        })
    return rows


def parse_csv(path):
    with open(path, 'r', encoding='utf-8-sig', newline='') as fh:
        sample = fh.read(4096); fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        rdr = csv.DictReader(open(path, 'r', encoding='utf-8-sig', newline=''), dialect=dialect)
        rows = []
        for r in rdr:
            # naive mapping
            rows.append({
                'gpstime': r.get('GPS时间') or r.get('gpstime') or r.get('time'),
                'lon': r.get('经度') or r.get('lon'),
                'lat': r.get('纬度') or r.get('lat'),
                'velocity': r.get('速度') or r.get('velocity'),
                'width': r.get('幅宽') or r.get('width'),
                'depth': r.get('深度') or r.get('depth'),
                'depthstandard': r.get('深度标准值') or r.get('depthstandard') or r.get('标准值'),
            })
        return rows


def to_float(v):
    if v is None or v == '':
        return None
    try:
        if isinstance(v, str):
            t = v.strip().replace('，', ',')
            if ',' in t and '.' not in t:
                t = t.replace(',', '.')
            return float(t)
        return float(v)
    except Exception:
        return None


def main(argv):
    if len(argv) < 3:
        print('Usage: fix_track_from_file.py <file_path> <trackid>')
        return 2
    file_path = Path(argv[1])
    trackid = int(argv[2])
    if not file_path.exists():
        print('file not found', file_path)
        return 2

    rows = parse_xlsx(file_path) if file_path.suffix.lower() in ('.xlsx', '.xls') else parse_csv(file_path)
    print('parsed rows', len(rows))

    db = Path(__file__).resolve().parents[2] / 'db.sqlite3'
    conn = sqlite3.connect(str(db), timeout=60)
    conn.execute('PRAGMA busy_timeout = 60000')
    cur = conn.cursor()

    # update by order: imported rows and stored points were created in the same sequence
    cur.execute('SELECT rowid FROM trackpoints WHERE trackid = ? ORDER BY rowid', (trackid,))
    db_ids = [row[0] for row in cur.fetchall()]
    update_params = []
    for db_id, r in zip(db_ids, rows):
        vel = to_float(r.get('velocity'))
        wid = to_float(r.get('width'))
        dep = to_float(r.get('depth'))
        dep_std = to_float(r.get('depthstandard'))
        update_params.append((vel, wid, dep, dep_std, db_id))

    updated = 0
    batch_size = 200
    updated = 0
    for start in range(0, len(update_params), batch_size):
        chunk = update_params[start:start + batch_size]
        cur.executemany(
            'UPDATE trackpoints SET velocity = COALESCE(?, velocity), width = COALESCE(?, width), depth = COALESCE(?, depth), depthstandard = COALESCE(?, depthstandard) WHERE rowid = ?',
            chunk,
        )
        conn.commit()
        updated += len(chunk)
        print(f'updated {updated}/{len(update_params)}', flush=True)

    conn.commit()

    # recompute track.width and work.avgvelocity/work.workarea
    cur.execute('SELECT AVG(width) FROM trackpoints WHERE trackid = ?', (trackid,))
    avg_w = cur.fetchone()[0] or 0.0
    cur.execute('SELECT SUM(CASE WHEN velocity IS NOT NULL THEN velocity ELSE 0 END), COUNT(*) FROM trackpoints WHERE trackid = ?', (trackid,))
    s = cur.fetchone()
    sum_v = s[0] or 0.0
    cnt = s[1] or 0
    avg_v = (sum_v / cnt) if cnt else 0.0

    cur.execute('UPDATE track SET width = ? WHERE trackid = ?', (avg_w, trackid))
    cur.execute('SELECT gpstime, lon, lat FROM trackpoints WHERE trackid = ? ORDER BY gpstime', (trackid,))
    pts = cur.fetchall()
    def hav(a,b):
        import math
        lon1, lat1 = math.radians(a[1]), math.radians(a[2])
        lon2, lat2 = math.radians(b[1]), math.radians(b[2])
        dlon = lon2-lon1; dlat = lat2-lat1
        return 6371000 * 2 * math.asin(math.sqrt(math.sin(dlat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2))
    total_m = 0.0
    for i in range(1, len(pts)):
        total_m += hav(pts[i-1], pts[i])

    work_area = (total_m * float(avg_w or 0.0)) / 10000.0
    cur.execute('UPDATE work SET workarea = ?, avgvelocity = ? WHERE trackid = ?', (work_area, avg_v, trackid))
    conn.commit()
    conn.close()

    print('updated rows', updated, 'avg_width', avg_w, 'avg_velocity', avg_v, 'work_area', work_area)
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv))
