#!/usr/bin/env python3
"""Backfill existing track rows from source XLSX files into the root sqlite database.

This script updates trackpoints by row order, then recomputes track/work/rate rows.
It is intended for the existing eight tracks already present in the database.
"""
from __future__ import annotations

import csv
import math
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

TRACK_IDS = [5, 6, 7, 8, 9, 12, 13, 14]


def normalize(text):
    import re
    return re.sub(r"[\s\-_/\\.:：，,。()（）\[\]{}]+", "", str(text).strip().lower())


def to_float(value):
    if value is None or value == "":
        return None
    try:
        if isinstance(value, str):
            text = value.strip().replace("，", ",")
            if "," in text and "." not in text:
                text = text.replace(",", ".")
            return float(text)
        return float(value)
    except Exception:
        return None


def to_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def parse_datetime_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return None


def parse_workbook(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = [normalize(value) for value in headers]

    index_map = {}
    for index, header in enumerate(headers):
        if header == "序列号" or header == "序号":
            continue
        if header in ("gpstime", "time", "时间", "gps时间") or "time" in header:
            index_map.setdefault("gpstime", index)
        elif "lon" in header or "经度" in header:
            index_map.setdefault("lon", index)
        elif "lat" in header or "纬度" in header:
            index_map.setdefault("lat", index)
        elif "speed" in header or "velocity" in header or "速度" in header:
            index_map.setdefault("velocity", index)
        elif "工作状态" in header or "workstatus" in header or "状态" in header:
            index_map.setdefault("workstatus", index)
        elif "width" in header or "幅宽" in header:
            index_map.setdefault("width", index)
        elif "深度标准" in header or "标准值" in header:
            index_map.setdefault("depthstandard", index)
        elif "深度" in header or "耕深" in header:
            index_map.setdefault("depth", index)
        elif header == "x":
            index_map.setdefault("x", index)
        elif header == "y":
            index_map.setdefault("y", index)
        elif "航向" in header or "course" in header:
            index_map.setdefault("course", index)

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        def get(field):
            idx = index_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        rows.append(
            {
                "gpstime": parse_datetime_value(get("gpstime")),
                "lon": to_float(get("lon")),
                "lat": to_float(get("lat")),
                "x": to_float(get("x")),
                "y": to_float(get("y")),
                "velocity": to_float(get("velocity")),
                "course": to_float(get("course")),
                "workstatus": to_int(get("workstatus")),
                "width": to_float(get("width")),
                "depth": to_float(get("depth")),
                "depthstandard": to_float(get("depthstandard")),
            }
        )
    return rows


def haversine(lat1, lon1, lat2, lon2):
    radius_km = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * radius_km * math.asin(math.sqrt(a))


def backfill_track(conn: sqlite3.Connection, root: Path, trackid: int):
    source = root / "BackEnd" / "datasets" / "xlsx" / f"{trackid}.xlsx"
    if not source.exists():
        print(f"skip track {trackid}: missing {source}")
        return False

    rows = parse_workbook(source)
    if not rows:
        print(f"skip track {trackid}: no parsed rows")
        return False

    cur = conn.cursor()
    cur.execute("SELECT rowid FROM trackpoints WHERE trackid = ? ORDER BY rowid", (trackid,))
    rowids = [row[0] for row in cur.fetchall()]
    if not rowids:
        print(f"skip track {trackid}: no existing trackpoints")
        return False

    updates = []
    for rowid, point in zip(rowids, rows):
        updates.append(
            (
                point["velocity"],
                point["workstatus"],
                point["width"],
                point["depth"],
                point["depthstandard"],
                rowid,
            )
        )

    batch_size = 500
    updated = 0
    for start in range(0, len(updates), batch_size):
        chunk = updates[start : start + batch_size]
        cur.executemany(
            "UPDATE trackpoints SET velocity = COALESCE(?, velocity), workstatus = COALESCE(?, workstatus), width = COALESCE(?, width), depth = COALESCE(?, depth), depthstandard = COALESCE(?, depthstandard) WHERE rowid = ?",
            chunk,
        )
        conn.commit()
        updated += len(chunk)
        print(f"track {trackid}: updated {updated}/{len(updates)}")

    cur.execute("SELECT AVG(width) FROM trackpoints WHERE trackid = ?", (trackid,))
    avg_width = cur.fetchone()[0] or 0.0

    cur.execute("SELECT gpstime, lon, lat, velocity, workstatus FROM trackpoints WHERE trackid = ? ORDER BY gpstime", (trackid,))
    point_rows = cur.fetchall()
    if not point_rows:
        print(f"track {trackid}: no points after update")
        return False

    start_time = point_rows[0][0]
    end_time = point_rows[-1][0]
    total_distance_m = 0.0
    for index in range(1, len(point_rows)):
        prev = point_rows[index - 1]
        curr = point_rows[index]
        if prev[1] is None or prev[2] is None or curr[1] is None or curr[2] is None:
            continue
        total_distance_m += haversine(prev[2], prev[1], curr[2], curr[1]) * 1000.0

    velocities = [float(row[3]) for row in point_rows if row[3] is not None]
    avg_velocity = round(sum(velocities) / len(velocities), 3) if velocities else 0.0

    workstatus_values = [int(row[4]) for row in point_rows if row[4] is not None]
    active_count = sum(1 for value in workstatus_values if value != 0)
    pass_rate = round((active_count / len(workstatus_values)) * 100.0, 2) if workstatus_values else 0.0
    work_time_hours = 0.0
    if start_time and end_time:
        work_time_hours = max((parse_datetime_value(end_time) - parse_datetime_value(start_time)).total_seconds(), 0) / 3600.0 if isinstance(start_time, str) or isinstance(end_time, str) else max((end_time - start_time).total_seconds(), 0) / 3600.0

    work_area = round((total_distance_m * float(avg_width or 0.0)) / 10000.0, 6)
    production_rate = round((work_area / work_time_hours), 3) if work_time_hours else 0.0

    cur.execute("UPDATE track SET starttime = ?, endtime = ?, width = ?, totalpoints = ? WHERE trackid = ?", (
        start_time,
        end_time,
        avg_width,
        len(point_rows),
        trackid,
    ))
    cur.execute("UPDATE work SET worktime = ?, worklength = ?, workarea = ?, avgvelocity = ? WHERE trackid = ?", (
        round(work_time_hours, 3),
        round(total_distance_m / 1000.0, 3),
        work_area,
        avg_velocity,
        trackid,
    ))
    cur.execute("REPLACE INTO rate (trackid, passrate, productionrate, timerrate) VALUES (?, ?, ?, ?)", (
        trackid,
        pass_rate,
        production_rate,
        pass_rate,
    ))
    conn.commit()

    print(f"track {trackid}: width={avg_width:.4f}, avg_velocity={avg_velocity:.3f}, pass_rate={pass_rate:.2f}, work_area={work_area:.6f}")
    return True


def main():
    root = Path(__file__).resolve().parents[2]
    db_path = root / "db.sqlite3"
    conn = sqlite3.connect(str(db_path), timeout=60)
    conn.execute("PRAGMA busy_timeout = 60000")

    try:
        for trackid in TRACK_IDS:
            backfill_track(conn, root, trackid)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
