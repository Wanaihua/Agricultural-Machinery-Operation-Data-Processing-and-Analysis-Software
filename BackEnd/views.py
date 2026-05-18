import hashlib
import csv
import math
import os
import re
import uuid
import time
from datetime import datetime
from statistics import mean
from pathlib import Path
from urllib.parse import urlparse

import pymysql
from django.conf import settings
from django.db import transaction
from django.http import JsonResponse
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response

from openpyxl import load_workbook
import xlrd

from BackEnd.generated_api.models import File as DatasetFile
from BackEnd.generated_api.models import ImportLog, Rate, Track, Trackpoints, Work


def ok(data=None, msg="success"):
    return Response({"code": "200", "data": data, "msg": msg})


def fail(msg="failed", code="500", data=None):
    return Response({"code": str(code), "data": data, "msg": msg})


def _db_conf():
    return {
        "host": os.getenv("MYSQL_HOST", "127.0.0.1"),
        "port": int(os.getenv("MYSQL_PORT", "3306")),
        "user": os.getenv("MYSQL_USER", "root"),
        "password": os.getenv("MYSQL_PASSWORD", "wanaihua"),
        "database": os.getenv("MYSQL_DB", "agricultural_machinery_db"),
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor,
        "autocommit": True,
    }


def _fetch_all(sql, params=None):
    conn = pymysql.connect(**_db_conf())
    try:
        with conn.cursor() as cursor:
            cursor.execute(sql, params or ())
            return cursor.fetchall()
    finally:
        conn.close()


def _fetch_one(sql, params=None):
    rows = _fetch_all(sql, params)
    return rows[0] if rows else None


def _execute(sql, params=None):
    conn = pymysql.connect(**_db_conf())
    try:
        with conn.cursor() as cursor:
            count = cursor.execute(sql, params or ())
            return count
    finally:
        conn.close()


def _build_menu_tree(rows):
    by_id = {}
    roots = []
    for item in rows:
        node = {
            "id": item.get("id"),
            "name": item.get("name"),
            "path": item.get("path"),
            "icon": item.get("icon"),
            "description": item.get("description"),
            "pid": item.get("pid"),
            "pagePath": item.get("pagePath"),
            "children": [],
        }
        by_id[node["id"]] = node

    for node in by_id.values():
        pid = node.get("pid")
        if pid and pid in by_id:
            by_id[pid]["children"].append(node)
        else:
            roots.append(node)

    return roots


def _to_md5(text):
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def _json_response(message="success", data=None, code=200, status=200):
    return JsonResponse({"code": code, "message": message, "data": data or {}}, status=status)


def _normalize_header(value):
    return re.sub(r"[\s\-_/\\.:：，,。()（）\[\]{}]+", "", str(value).strip().lower())


def _parse_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _parse_datetime_value(value, datemode=0):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return xlrd.xldate_as_datetime(value, datemode)
        except Exception:
            return None

    text = str(value).strip()
    patterns = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _resolve_dataset_file(file_url_or_path):
    if not file_url_or_path:
        return None

    raw_path = str(file_url_or_path).replace("\\", "/")
    parsed_path = urlparse(raw_path).path if raw_path.startswith(("http://", "https://")) else raw_path
    parsed_path = parsed_path.lstrip("/")

    if parsed_path.startswith("datasets/"):
        return Path(settings.BASE_DIR) / parsed_path

    return Path(settings.DATASETS_ROOT) / parsed_path


def _haversine_distance_meters(point_a, point_b):
    if point_a is None or point_b is None:
        return 0.0

    lon1, lat1 = math.radians(float(point_a["lon"])), math.radians(float(point_a["lat"]))
    lon2, lat2 = math.radians(float(point_b["lon"])), math.radians(float(point_b["lat"]))
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 6371000 * 2 * math.asin(math.sqrt(a))


def _build_column_map(header_row):
    aliases = {
        "gpstime": ["gpstime", "time", "timestamp", "gps时间", "时间", "定位时间", "采集时间"],
        "lon": ["lon", "longitude", "经度"],
        "lat": ["lat", "latitude", "纬度"],
        "x": ["x"],
        "y": ["y"],
        "velocity": ["velocity", "speed", "速度"],
        "course": ["course", "航向", "方向角"],
        "workstatus": ["workstatus", "status", "作业状态", "状态"],
        "width": ["width", "幅宽", "作业幅宽"],
        "depth": ["depth", "plowingdepth", "耕深", "深度"],
        "depthstandard": ["depthstandard", "标准耕深", "耕深标准", "标准值"],
    }

    column_map = {}
    normalized_headers = {_normalize_header(value): index for index, value in enumerate(header_row)}
    for field_name, candidates in aliases.items():
        for candidate in candidates:
            candidate_key = _normalize_header(candidate)
            if candidate_key in normalized_headers:
                column_map[field_name] = normalized_headers[candidate_key]
                break
    return column_map


def _parse_rows(rows, datemode=0):
    parsed_rows = []
    errors = []

    for index, row in enumerate(rows, start=2):
        point = {
            "gpstime": None,
            "lon": None,
            "lat": None,
            "x": None,
            "y": None,
            "velocity": None,
            "course": None,
            "workstatus": None,
            "width": None,
            "depth": None,
            "depthstandard": None,
        }

        try:
            if isinstance(row, dict):
                row_map = { _normalize_header(key): value for key, value in row.items() }
                for field_name, aliases in {
                    "gpstime": ["gpstime", "time", "timestamp", "gps时间", "时间", "定位时间", "采集时间"],
                    "lon": ["lon", "longitude", "经度"],
                    "lat": ["lat", "latitude", "纬度"],
                    "x": ["x"],
                    "y": ["y"],
                    "velocity": ["velocity", "speed", "速度"],
                    "course": ["course", "航向", "方向角"],
                    "workstatus": ["workstatus", "status", "作业状态", "状态"],
                    "width": ["width", "幅宽", "作业幅宽"],
                    "depth": ["depth", "plowingdepth", "耕深", "深度"],
                    "depthstandard": ["depthstandard", "标准耕深", "耕深标准", "标准值"],
                }.items():
                    for alias in aliases:
                        alias_key = _normalize_header(alias)
                        if alias_key in row_map:
                            point[field_name] = row_map[alias_key]
                            break
            else:
                header_map = row[0]
                values = row[1]
                for field_name, column_index in header_map.items():
                    if 0 <= column_index < len(values):
                        point[field_name] = values[column_index]

            point["gpstime"] = _parse_datetime_value(point["gpstime"], datemode)
            point["lon"] = _parse_float(point["lon"])
            point["lat"] = _parse_float(point["lat"])
            point["x"] = _parse_float(point["x"])
            point["y"] = _parse_float(point["y"])
            point["velocity"] = _parse_float(point["velocity"])
            point["course"] = _parse_float(point["course"])
            point["workstatus"] = _parse_int(point["workstatus"])
            point["width"] = _parse_float(point["width"])
            point["depth"] = _parse_float(point["depth"])
            point["depthstandard"] = _parse_float(point["depthstandard"])

            if point["lon"] is not None and not (-180 <= point["lon"] <= 180):
                errors.append(f"第{index}行经度异常，已跳过")
                continue
            if point["lat"] is not None and not (-90 <= point["lat"] <= 90):
                errors.append(f"第{index}行纬度异常，已跳过")
                continue
            if point["gpstime"] is None or point["lon"] is None or point["lat"] is None:
                errors.append(f"第{index}行关键字段缺失，已跳过")
                continue

            parsed_rows.append(point)
        except Exception as exc:
            errors.append(f"第{index}行解析失败: {exc}")

    return parsed_rows, errors


def parse_csv(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        except Exception:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        return _parse_rows(reader)


def parse_excel(file_path):
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xls":
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        headers = [_normalize_header(sheet.cell_value(0, col_index)) for col_index in range(sheet.ncols)]
        column_map = _build_column_map(headers)
        rows = []
        for row_index in range(1, sheet.nrows):
            row_values = [sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)]
            row_dict = {}
            for field_name, column_index in column_map.items():
                row_dict[field_name] = row_values[column_index] if column_index < len(row_values) else None
            rows.append(row_dict)
        return _parse_rows(rows, datemode=workbook.datemode)

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    column_map = _build_column_map(header_values)
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for field_name, column_index in column_map.items():
            row_dict[field_name] = row[column_index] if column_index < len(row) else None
        rows.append(row_dict)
    return _parse_rows(rows)


def _parse_track_file(file_path):
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        return parse_csv(file_path)
    if suffix in {".xlsx", ".xls"}:
        return parse_excel(file_path)
    raise ValueError("仅支持 CSV / Excel 文件解析")


def _save_import_log(admin_id, file_name, import_count, import_status, error_info):
    return ImportLog.objects.create(
        admin_id_id=admin_id,
        file_name=file_name,
        import_count=import_count,
        import_status=import_status,
        error_info=error_info or None,
        import_time=timezone.now(),
    )


_USER_HAS_IS_DELETE = None


def _user_has_is_delete():
    global _USER_HAS_IS_DELETE
    if _USER_HAS_IS_DELETE is not None:
        return _USER_HAS_IS_DELETE

    row = _fetch_one(
        """
        SELECT COUNT(*) AS cnt
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = 'user' AND column_name = 'is_delete'
        """,
        (_db_conf()["database"],),
    )
    _USER_HAS_IS_DELETE = bool(row and row.get("cnt", 0) > 0)
    return _USER_HAS_IS_DELETE
 
@api_view(['GET'])
def get_data(request):
    data = {'message': 'Hello from Django!'}
    return Response(data)


@api_view(['POST'])
def login_test(request):
    username = request.data.get('username')
    password = request.data.get('password')
    db_host = os.getenv('MYSQL_HOST', '127.0.0.1')
    db_port = int(os.getenv('MYSQL_PORT', '3306'))
    db_user = os.getenv('MYSQL_USER', 'root')
    db_password = os.getenv('MYSQL_PASSWORD', 'wanaihua')
    db_name = os.getenv('MYSQL_DB', 'agricultural_machinery_db')

    conn = None

    try:
        conn = pymysql.connect(
            host=db_host,
            port=db_port,
            user=db_user,
            password=db_password,
            database=db_name,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
        )
        with conn.cursor() as cursor:
            cursor.execute('SELECT 1')
    except Exception as exc:
        return Response({'ok': False, 'connected': False, 'error': str(exc)}, status=500)
    finally:
        if conn is not None:
            conn.close()

    if not username or not password:
        return Response({'ok': False, 'connected': True, 'message': 'username/password required'}, status=400)

    md5_password = hashlib.md5(password.encode('utf-8')).hexdigest()

    try:
        conn = pymysql.connect(
            host=db_host,
            port=db_port,
            user=db_user,
            password=db_password,
            database=db_name,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
        )
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, role
                FROM user_info
                WHERE username = %s AND LOWER(password) = %s AND IFNULL(is_delete, 0) = 0
                LIMIT 1
                """,
                (username, md5_password),
            )
            user = cursor.fetchone()
    except Exception as exc:
        return Response({'ok': False, 'connected': False, 'error': str(exc)}, status=500)
    finally:
        if conn is not None:
            conn.close()

    if user:
        return Response({'ok': True, 'connected': True, 'user_id': user['id'], 'role': user['role']})

    return Response({'ok': False, 'connected': True, 'message': 'invalid credentials'}, status=401)


@api_view(["POST"])
def user_login(request):
    username = (request.data.get("username") or "").strip()
    password = request.data.get("password") or ""
    if not username or not password:
        return fail("用户名或密码不能为空", code="400")

    md5_password = _to_md5(password)
    active_filter = "AND IFNULL(is_delete, 0) = 0" if _user_has_is_delete() else ""
    sql = f"""
    SELECT id, username, nickname, email, phone, address, avatar_url AS avatarUrl, role
    FROM user
    WHERE username = %s
      {active_filter}
      AND (LOWER(password) = %s OR password = %s)
    LIMIT 1
    """
    user = _fetch_one(sql, (username, md5_password, password))
    if not user:
        return fail("用户名或密码错误", code="401")

    role_flag = user.get("role")
    menus_sql = """
    SELECT m.id, m.name, m.path, m.icon, m.description, m.pid, m.page_path AS pagePath
    FROM menu m
    JOIN role_menu rm ON rm.menu_id = m.id
    JOIN role r ON r.id = rm.role_id
    WHERE r.flag = %s
    ORDER BY m.id
    """
    menus = _fetch_all(menus_sql, (role_flag,)) if role_flag else []
    user["menus"] = _build_menu_tree(menus)
    user["token"] = str(uuid.uuid4())
    return ok(user, "登录成功")


@api_view(["POST"])
def user_register(request):
    username = (request.data.get("username") or "").strip()
    password = request.data.get("password") or ""
    if not username or not password:
        return fail("用户名或密码不能为空", code="400")

    exists = _fetch_one("SELECT id FROM user WHERE username = %s LIMIT 1", (username,))
    if exists:
        return fail("用户名已存在", code="400")

    if _user_has_is_delete():
        _execute(
            """
            INSERT INTO user(username, password, nickname, role, is_delete)
            VALUES(%s, %s, %s, %s, 0)
            """,
            (username, _to_md5(password), username, "ROLE_USER"),
        )
    else:
        _execute(
            """
            INSERT INTO user(username, password, nickname, role)
            VALUES(%s, %s, %s, %s)
            """,
            (username, _to_md5(password), username, "ROLE_USER"),
        )
    user = _fetch_one(
        "SELECT id, username, nickname, email, phone, address, avatar_url AS avatarUrl, role FROM user WHERE username=%s LIMIT 1",
        (username,),
    )
    user["menus"] = []
    user["token"] = str(uuid.uuid4())
    return ok(user, "注册成功")


@api_view(["GET"])
def user_by_username(request, username):
    user = _fetch_one(
        """
        SELECT id, username, password, nickname, email, phone, address, avatar_url AS avatarUrl, role
        FROM user
        WHERE username = %s
        LIMIT 1
        """,
        (username,),
    )
    return ok(user)


@api_view(["GET"])
def user_page(request):
    page_num = int(request.query_params.get("pageNum", 1))
    page_size = int(request.query_params.get("pageSize", 10))
    username = request.query_params.get("username", "")
    address = request.query_params.get("address", "")
    email = request.query_params.get("email", "")
    active_where = "IFNULL(is_delete,0)=0 AND " if _user_has_is_delete() else ""
    where = f"WHERE {active_where}username LIKE %s AND address LIKE %s AND email LIKE %s"
    params = (f"%{username}%", f"%{address}%", f"%{email}%")
    total_row = _fetch_one(f"SELECT COUNT(*) AS total FROM user {where}", params)
    total = total_row["total"] if total_row else 0
    offset = (page_num - 1) * page_size
    rows = _fetch_all(
        f"""
        SELECT id, username, nickname, email, phone, address, avatar_url AS avatarUrl, role
        FROM user
        {where}
        ORDER BY id DESC
        LIMIT %s OFFSET %s
        """,
        params + (page_size, offset),
    )
    return ok({"records": rows, "total": total})


@api_view(["POST"])
def user_save(request):
    body = request.data or {}
    user_id = body.get("id")
    password = body.get("password")
    if password and len(password) != 32:
        body["password"] = _to_md5(password)

    if user_id:
        fields = []
        values = []
        for key, col in [
            ("username", "username"),
            ("password", "password"),
            ("nickname", "nickname"),
            ("email", "email"),
            ("phone", "phone"),
            ("address", "address"),
            ("avatarUrl", "avatar_url"),
            ("role", "role"),
        ]:
            if key in body and body.get(key) is not None:
                fields.append(f"{col}=%s")
                values.append(body.get(key))
        if fields:
            values.append(user_id)
            _execute(f"UPDATE user SET {', '.join(fields)} WHERE id=%s", tuple(values))
        return ok(True, "保存成功")

    if _user_has_is_delete():
        _execute(
            """
            INSERT INTO user(username, password, nickname, email, phone, address, avatar_url, role, is_delete)
            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, 0)
            """,
            (
                body.get("username"),
                body.get("password") or _to_md5("123456"),
                body.get("nickname"),
                body.get("email"),
                body.get("phone"),
                body.get("address"),
                body.get("avatarUrl"),
                body.get("role") or "ROLE_USER",
            ),
        )
    else:
        _execute(
            """
            INSERT INTO user(username, password, nickname, email, phone, address, avatar_url, role)
            VALUES(%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                body.get("username"),
                body.get("password") or _to_md5("123456"),
                body.get("nickname"),
                body.get("email"),
                body.get("phone"),
                body.get("address"),
                body.get("avatarUrl"),
                body.get("role") or "ROLE_USER",
            ),
        )
    return ok(True, "新增成功")


@api_view(["POST"])
def user_update_password(request):
    username = (request.data.get("username") or "").strip()
    old_password = request.data.get("oldPassword") or ""
    new_password = request.data.get("newPassword") or ""

    if not username or not old_password or not new_password:
        return fail("参数不完整", code="400")

    active_filter = "AND IFNULL(is_delete,0)=0" if _user_has_is_delete() else ""
    user = _fetch_one(
        f"SELECT id, password FROM user WHERE username=%s {active_filter} LIMIT 1",
        (username,),
    )
    if not user:
        return fail("用户不存在", code="404")

    stored_password = (user.get("password") or "").lower()
    old_md5 = _to_md5(old_password)
    # Compatible with both legacy plaintext and md5 storage.
    if stored_password not in {old_password.lower(), old_md5.lower()}:
        return fail("原密码错误", code="400")

    _execute("UPDATE user SET password=%s WHERE id=%s", (_to_md5(new_password), user["id"]))
    return ok(True, "密码修改成功")


@api_view(["DELETE"])
def user_delete(request, user_id):
    if _user_has_is_delete():
        _execute("UPDATE user SET is_delete=1 WHERE id=%s", (user_id,))
    else:
        _execute("DELETE FROM user WHERE id=%s", (user_id,))
    return ok(True, "删除成功")


@api_view(["POST"])
def user_delete_batch(request):
    ids = request.data or []
    if not ids:
        return fail("未提供ID", code="400")
    placeholders = ",".join(["%s"] * len(ids))
    if _user_has_is_delete():
        _execute(f"UPDATE user SET is_delete=1 WHERE id IN ({placeholders})", tuple(ids))
    else:
        _execute(f"DELETE FROM user WHERE id IN ({placeholders})", tuple(ids))
    return ok(True, "批量删除成功")


@api_view(["GET"])
def role_list(request):
    rows = _fetch_all("SELECT id, name, flag, description FROM role ORDER BY id DESC")
    return ok(rows)


@api_view(["GET", "POST"])
def role_entry(request):
    if request.method == "GET":
        rows = _fetch_all("SELECT id, name, flag, description FROM role ORDER BY id DESC")
        return ok(rows)

    body = request.data or {}
    role_id = body.get("id")
    if role_id:
        _execute(
            "UPDATE role SET name=%s, flag=%s, description=%s WHERE id=%s",
            (body.get("name"), body.get("flag"), body.get("description"), role_id),
        )
        return ok(True, "保存成功")

    _execute(
        "INSERT INTO role(name, flag, description) VALUES(%s, %s, %s)",
        (body.get("name"), body.get("flag"), body.get("description")),
    )
    return ok(True, "新增成功")


@api_view(["GET"])
def role_page(request):
    page_num = int(request.query_params.get("pageNum", 1))
    page_size = int(request.query_params.get("pageSize", 10))
    name = request.query_params.get("name", "")
    total_row = _fetch_one("SELECT COUNT(*) AS total FROM role WHERE name LIKE %s", (f"%{name}%",))
    total = total_row["total"] if total_row else 0
    offset = (page_num - 1) * page_size
    rows = _fetch_all(
        "SELECT id, name, flag, description FROM role WHERE name LIKE %s ORDER BY id DESC LIMIT %s OFFSET %s",
        (f"%{name}%", page_size, offset),
    )
    return ok({"records": rows, "total": total})


@api_view(["POST"])
def role_save(request):
    body = request.data or {}
    role_id = body.get("id")
    if role_id:
        _execute(
            "UPDATE role SET name=%s, flag=%s, description=%s WHERE id=%s",
            (body.get("name"), body.get("flag"), body.get("description"), role_id),
        )
        return ok(True, "保存成功")
    _execute(
        "INSERT INTO role(name, flag, description) VALUES(%s, %s, %s)",
        (body.get("name"), body.get("flag"), body.get("description")),
    )
    return ok(True, "新增成功")


@api_view(["DELETE"])
def role_delete(request, role_id):
    _execute("DELETE FROM role_menu WHERE role_id=%s", (role_id,))
    _execute("DELETE FROM role WHERE id=%s", (role_id,))
    return ok(True, "删除成功")


@api_view(["POST"])
def role_delete_batch(request):
    ids = request.data or []
    if not ids:
        return fail("未提供ID", code="400")
    placeholders = ",".join(["%s"] * len(ids))
    _execute(f"DELETE FROM role_menu WHERE role_id IN ({placeholders})", tuple(ids))
    _execute(f"DELETE FROM role WHERE id IN ({placeholders})", tuple(ids))
    return ok(True, "批量删除成功")


@api_view(["POST"])
def role_role_menu_save(request, role_id):
    menu_ids = request.data or []
    _execute("DELETE FROM role_menu WHERE role_id=%s", (role_id,))
    for menu_id in menu_ids:
        _execute("INSERT INTO role_menu(role_id, menu_id) VALUES(%s, %s)", (role_id, menu_id))
    return ok(True, "保存成功")


@api_view(["GET", "POST"])
def role_role_menu_entry(request, role_id):
    if request.method == "GET":
        rows = _fetch_all("SELECT menu_id FROM role_menu WHERE role_id=%s", (role_id,))
        return ok([row["menu_id"] for row in rows])

    menu_ids = request.data or []
    _execute("DELETE FROM role_menu WHERE role_id=%s", (role_id,))
    for menu_id in menu_ids:
        _execute("INSERT INTO role_menu(role_id, menu_id) VALUES(%s, %s)", (role_id, menu_id))
    return ok(True, "保存成功")


@api_view(["GET"])
def role_role_menu(request, role_id):
    rows = _fetch_all("SELECT menu_id FROM role_menu WHERE role_id=%s", (role_id,))
    return ok([row["menu_id"] for row in rows])


@api_view(["GET"])
def menu_list(request):
    name = request.query_params.get("name", "")
    rows = _fetch_all(
        """
        SELECT id, name, path, icon, description, pid, page_path AS pagePath
        FROM menu
        WHERE name LIKE %s
        ORDER BY id
        """,
        (f"%{name}%",),
    )
    return ok(_build_menu_tree(rows))


@api_view(["GET", "POST"])
def menu_entry(request):
    if request.method == "GET":
        name = request.query_params.get("name", "")
        rows = _fetch_all(
            """
            SELECT id, name, path, icon, description, pid, page_path AS pagePath
            FROM menu
            WHERE name LIKE %s
            ORDER BY id
            """,
            (f"%{name}%",),
        )
        return ok(_build_menu_tree(rows))

    body = request.data or {}
    menu_id = body.get("id")
    if menu_id:
        _execute(
            "UPDATE menu SET name=%s, path=%s, icon=%s, description=%s, pid=%s, page_path=%s WHERE id=%s",
            (
                body.get("name"),
                body.get("path"),
                body.get("icon"),
                body.get("description"),
                body.get("pid"),
                body.get("pagePath"),
                menu_id,
            ),
        )
        return ok(True, "保存成功")

    _execute(
        "INSERT INTO menu(name, path, icon, description, pid, page_path) VALUES(%s, %s, %s, %s, %s, %s)",
        (
            body.get("name"),
            body.get("path"),
            body.get("icon"),
            body.get("description"),
            body.get("pid"),
            body.get("pagePath"),
        ),
    )
    return ok(True, "新增成功")


@api_view(["GET"])
def menu_ids(request):
    rows = _fetch_all("SELECT id FROM menu ORDER BY id")
    return ok([r["id"] for r in rows])


@api_view(["GET"])
def menu_icons(request):
    icons = [
        {"name": "用户", "value": "el-icon-user"},
        {"name": "菜单", "value": "el-icon-menu"},
        {"name": "设置", "value": "el-icon-setting"},
        {"name": "数据", "value": "el-icon-data-analysis"},
        {"name": "文档", "value": "el-icon-document"},
    ]
    return ok(icons)


@api_view(["POST"])
def menu_save(request):
    body = request.data or {}
    menu_id = body.get("id")
    if menu_id:
        _execute(
            "UPDATE menu SET name=%s, path=%s, icon=%s, description=%s, pid=%s, page_path=%s WHERE id=%s",
            (
                body.get("name"),
                body.get("path"),
                body.get("icon"),
                body.get("description"),
                body.get("pid"),
                body.get("pagePath"),
                menu_id,
            ),
        )
        return ok(True, "保存成功")
    _execute(
        "INSERT INTO menu(name, path, icon, description, pid, page_path) VALUES(%s, %s, %s, %s, %s, %s)",
        (
            body.get("name"),
            body.get("path"),
            body.get("icon"),
            body.get("description"),
            body.get("pid"),
            body.get("pagePath"),
        ),
    )
    return ok(True, "新增成功")


@api_view(["DELETE"])
def menu_delete(request, menu_id):
    _execute("DELETE FROM role_menu WHERE menu_id=%s", (menu_id,))
    _execute("DELETE FROM menu WHERE id=%s OR pid=%s", (menu_id, menu_id))
    return ok(True, "删除成功")


@api_view(["POST"])
def menu_delete_batch(request):
    ids = request.data or []
    if not ids:
        return fail("未提供ID", code="400")
    placeholders = ",".join(["%s"] * len(ids))
    _execute(f"DELETE FROM role_menu WHERE menu_id IN ({placeholders})", tuple(ids))
    _execute(f"DELETE FROM menu WHERE id IN ({placeholders})", tuple(ids))
    return ok(True, "批量删除成功")


@api_view(["GET"])
def file_page(request):
    page_num = int(request.query_params.get("pageNum", 1))
    page_size = int(request.query_params.get("pageSize", 10))
    name = request.query_params.get("name", "")
    total_row = _fetch_one("SELECT COUNT(*) AS total FROM file WHERE IFNULL(is_delete,0)=0 AND name LIKE %s", (f"%{name}%",))
    total = total_row["total"] if total_row else 0
    offset = (page_num - 1) * page_size
    rows = _fetch_all(
        """
        SELECT id, name, type, size, url, is_delete, enable, md5
        FROM file
        WHERE IFNULL(is_delete,0)=0 AND name LIKE %s
        ORDER BY id DESC
        LIMIT %s OFFSET %s
        """,
        (f"%{name}%", page_size, offset),
    )
    return ok({"records": rows, "total": total})


@api_view(["DELETE"])
def file_delete(request, file_id):
    _execute("UPDATE file SET is_delete=1 WHERE id=%s", (file_id,))
    return ok(True, "删除成功")


@api_view(["POST"])
def file_delete_batch(request):
    ids = request.data or []
    if not ids:
        return fail("未提供ID", code="400")
    placeholders = ",".join(["%s"] * len(ids))
    _execute(f"UPDATE file SET is_delete=1 WHERE id IN ({placeholders})", tuple(ids))
    return ok(True, "批量删除成功")


@api_view(["POST"])
def file_update(request):
    row = request.data or {}
    _execute(
        "UPDATE file SET name=%s, type=%s, size=%s, url=%s, enable=%s, md5=%s WHERE id=%s",
        (
            row.get("name"),
            row.get("type"),
            row.get("size"),
            row.get("url"),
            1 if row.get("enable") else 0,
            row.get("md5"),
            row.get("id"),
        ),
    )
    return ok(True, "操作成功")


@api_view(["POST"])
def upload_file(request):
    """Upload a single file and store it under BackEnd/datasets by type."""
    uploaded_file = request.FILES.get("file") or next(iter(request.FILES.values()), None)
    if uploaded_file is None:
        return _json_response("未获取到上传文件", code=400, status=400)

    suffix = Path(uploaded_file.name).suffix.lower()
    folder_map = {
        ".xlsx": "xlsx",
        ".xls": "xlsx",
        ".csv": "csv",
        ".png": "遥感图",
        ".jpg": "遥感图",
        ".jpeg": "遥感图",
    }
    target_folder = folder_map.get(suffix)
    if not target_folder:
        return _json_response("不支持的文件类型", code=400, status=400)

    target_dir = Path(settings.DATASETS_ROOT) / target_folder
    target_dir.mkdir(parents=True, exist_ok=True)

    stored_name = f"{uuid.uuid4().hex}{suffix}"
    file_path = target_dir / stored_name
    md5_ctx = hashlib.md5()
    total_bytes = 0

    with open(file_path, "wb") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)
            md5_ctx.update(chunk)
            total_bytes += len(chunk)

    relative_url = f"{settings.DATASETS_URL}{target_folder}/{stored_name}".replace("\\", "/")
    file_record = DatasetFile.objects.create(
        name=Path(uploaded_file.name).stem,
        type=suffix.lstrip(".").lower(),
        size=round(total_bytes / 1024, 2),
        url=relative_url,
        is_delete=False,
        enable=True,
        md5=md5_ctx.hexdigest(),
    )

    response_data = {
        "id": file_record.id,
        "name": file_record.name,
        "url": request.build_absolute_uri(relative_url),
        "size": file_record.size,
        "type": file_record.type,
        "md5": file_record.md5,
    }
    return _json_response(data=response_data)


@api_view(["POST"])
def import_data(request):
    """Import trajectory rows from an uploaded CSV / Excel file by file_id."""
    file_id = request.data.get("file_id") or request.data.get("fileId") or request.data.get("id")
    admin_id = request.data.get("admin_id") or request.data.get("adminId") or request.data.get("user_id") or 1

    if not file_id:
        return _json_response("缺少文件ID", code=400, status=400)

    file_record = DatasetFile.objects.filter(id=file_id, is_delete=False).first()
    if not file_record:
        return _json_response("文件不存在或已删除", code=404, status=404)

    file_path = _resolve_dataset_file(file_record.url)
    if file_path is None or not file_path.exists():
        return _json_response("文件路径不存在", code=404, status=404)

    start_time = time.perf_counter()
    parsed_rows, parse_errors = _parse_track_file(file_path)

    if not parsed_rows:
        error_text = "\n".join(parse_errors) if parse_errors else "未解析到有效数据"
        _save_import_log(admin_id, file_record.name, 0, "fail", error_text)
        return _json_response(
            message="导入失败",
            data={
                "success_count": 0,
                "failure_count": 0,
                "cost_seconds": round(time.perf_counter() - start_time, 3),
                "errors": parse_errors,
            },
            code=500,
            status=500,
        )

    valid_points = []
    for point in parsed_rows:
        valid_points.append(point)

    first_point = valid_points[0]
    last_point = valid_points[-1]
    track_width = next((item["width"] for item in valid_points if item.get("width") is not None), None)
    if track_width is None:
        track_width = 0.0

    total_distance_m = 0.0
    for index in range(1, len(valid_points)):
        total_distance_m += _haversine_distance_meters(valid_points[index - 1], valid_points[index])

    work_time_hours = 0.0
    if first_point.get("gpstime") and last_point.get("gpstime"):
        work_time_hours = max((last_point["gpstime"] - first_point["gpstime"]).total_seconds(), 0) / 3600.0

    avg_velocity = round(mean([item["velocity"] for item in valid_points if item.get("velocity") is not None]) if any(item.get("velocity") is not None for item in valid_points) else 0.0, 3)
    work_area = round((total_distance_m * float(track_width or 0.0)) / 10000.0, 3)
    active_points = [item for item in valid_points if (item.get("workstatus") or 0) != 0]
    pass_rate = round((len(active_points) / len(valid_points)) * 100.0, 2) if valid_points else 0.0
    production_rate = round((work_area / work_time_hours), 3) if work_time_hours else 0.0
    time_rate = round((len(active_points) / len(valid_points)) * 100.0, 2) if valid_points else 0.0

    error_info = "\n".join(parse_errors) if parse_errors else None

    try:
        with transaction.atomic():
            track = Track.objects.create(
                starttime=first_point["gpstime"],
                endtime=last_point["gpstime"],
                width=float(track_width or 0.0),
                totalpoints=len(valid_points),
            )

            trackpoints_objects = [
                Trackpoints(
                    trackid=track,
                    gpstime=item["gpstime"],
                    lon=item["lon"],
                    lat=item["lat"],
                    x=item.get("x"),
                    y=item.get("y"),
                    velocity=item.get("velocity"),
                    course=item.get("course"),
                    workstatus=item.get("workstatus"),
                    width=item.get("width"),
                    depth=item.get("depth"),
                    depthstandard=item.get("depthstandard"),
                )
                for item in valid_points
            ]
            Trackpoints.objects.bulk_create(trackpoints_objects, batch_size=500)

            Work.objects.create(
                trackid=track,
                worktime=round(work_time_hours, 3),
                worklength=round(total_distance_m / 1000.0, 3),
                workarea=work_area,
                avgvelocity=avg_velocity,
            )
            Rate.objects.create(
                trackid=track,
                passrate=pass_rate,
                productionrate=production_rate,
                timerrate=time_rate,
            )

            import_log = _save_import_log(
                admin_id,
                file_record.name,
                len(valid_points),
                "success",
                error_info,
            )
    except Exception as exc:
        import_log = _save_import_log(
            admin_id,
            file_record.name,
            len(valid_points),
            "fail",
            f"数据库写入失败: {exc}\n{error_info or ''}".strip(),
        )
        return _json_response(
            message="导入失败",
            data={
                "import_log_id": import_log.id,
                "success_count": 0,
                "failure_count": len(parsed_rows),
                "cost_seconds": round(time.perf_counter() - start_time, 3),
                "errors": parse_errors + [str(exc)],
            },
            code=500,
            status=500,
        )

    cost_seconds = round(time.perf_counter() - start_time, 3)
    return _json_response(
        data={
            "track_id": track.trackid,
            "import_log_id": import_log.id,
            "success_count": len(valid_points),
            "failure_count": len(parse_errors),
            "cost_seconds": cost_seconds,
            "errors": parse_errors,
        },
    )


# Backward-compatible alias for existing routes and callers.
file_upload = upload_file


@api_view(["GET"])
def echarts_members(request):
    rows = _fetch_all(
        """
        SELECT MONTH(creat_time) AS m, COUNT(*) AS c
        FROM user
        WHERE creat_time IS NOT NULL
        GROUP BY MONTH(creat_time)
        ORDER BY MONTH(creat_time)
        """
    )
    data = [0, 0, 0, 0]
    for row in rows:
        month = int(row.get("m") or 0)
        if 1 <= month <= 3:
            data[0] += row["c"]
        elif 4 <= month <= 6:
            data[1] += row["c"]
        elif 7 <= month <= 9:
            data[2] += row["c"]
        elif 10 <= month <= 12:
            data[3] += row["c"]
    return ok(data)